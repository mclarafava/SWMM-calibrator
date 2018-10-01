# -*- coding: utf-8 -*-
import openpyxl
import gc
import os
import numpy as np
from swmm5.swmm5tools import SWMM5Simulation

import xlrd
from datetime import datetime, timedelta
from collections import OrderedDict
from _consts import *
from parse_swmm import *
import dateutil.parser
from wadi_tools import *
from time import sleep
import glob
import sys

class SwmmTools:
    def __init__ (self, swmm_filename, data_filename = None, working_dir = None, temp_file = True):
        self.swmm_filename = swmm_filename
        self.data_filename = data_filename
        self.working_dir = working_dir
        self.root_dir = os.path.dirname(os.path.realpath(__file__))
        self.temp_file = temp_file
        self.st = None

    # run a simulation if necessary
    def check_simulation(self):
        if self.st == None:
            print('use run_swmm() first.')

    # Careful! this function round the start_date_time to midnight
    def get_start_date(self):
        self.check_simulation()
        dia = self.st.SWMM_StartDate
        seconds = (dia - 25569) * 86400
        return datetime.utcfromtimestamp(seconds).replace(minute=0, hour =0)

    def get_report_step(self):
        self.check_simulation()
        return self.st.SWMM_ReportStep

    def get_Nperiods(self):
        self.check_simulation()
        return self.st.SWMM_Nperiods

    def get_report_start_date(self):
        tempnet = ParseSwmm(self.swmm_filename)
        _, dia = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_DATE')
        _, hora = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_TIME')
        dia = dia.split()[1]
        hora = hora.split()[1]
        diahora = dia + ' ' + hora
        return dateutil.parser.parse(diahora)

    def get_report_start_date(self):
        tempnet = ParseSwmm(self.swmm_filename)
        _, dia = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_DATE')
        _, hora = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_TIME')
        dia = dia.split()[1]
        hora = hora.split()[1]
        diahora = dia + ' ' + hora
        return dateutil.parser.parse(diahora)

    # get the observed id from spreadsheet
    def get_observed_nodes (self):
        if self.data_filename == None:
            print('Function available only if a spreadsheet is defined.')
            return
        os.chdir(self.root_dir)
        wb = openpyxl.load_workbook(self.data_filename)
        return wb.sheetnames

    # get the simulated water depth by a specific id
    def get_level_by_id(self, id):
        self.check_simulation()
        return np.array(list(self.st.Results('NODE', id, depth_idx)))

    # get the simulated flow by a specific id
    def get_flow_by_id(self, id):
        self.check_simulation()
        return np.array(list(self.st.Results('NODE', id, flow_idx)))

    # recover simulated water depth from all POI
    def load_simulation_data(self):
        self.run_swmm()
        simulado_todos = np.array([],dtype=np.float)
        for id in self.get_observed_nodes():
            nivel_simulado = self.get_level_by_id(id)
            simulado_todos = np.concatenate ((simulado_todos, nivel_simulado))
        return simulado_todos

    def load_field_data(self):
        os.chdir(self.root_dir)
        wb = openpyxl.load_workbook(self.data_filename)
        observado_todos = np.array([],dtype=np.float)
        for aba in self.get_observed_nodes():
            sheet = wb[aba]
            nivel_medido = np.array([], dtype=np.float)
            for i in range(2, sheet.max_row+1):
                nivel_medido = np.append(nivel_medido, sheet.cell(row=i,column=3).value)
            observado_todos = np.concatenate ((observado_todos, nivel_medido))
        observado_todos = np.array(observado_todos, dtype=np.float)
        return observado_todos

    # average nash-sutcliffe efficiency (AVNSE)
    def calc_avnse(self, sim_data, field_data):
        ndata = self.st.SWMM_Nperiods
        yobs = field_data.reshape(int(len(field_data)/ndata), ndata)
        ysim = sim_data.reshape(int(len(sim_data)/ndata), ndata)
        nash_dict = {}
        wb = openpyxl.load_workbook(self.data_filename)
        abas = wb.sheetnames
        negative_count = 0
        nash_sum = 0
        n = len(abas)
        for i,aba in enumerate(abas):
            num = np.nansum((yobs[i] - ysim[i])**2)
            den = np.nansum((yobs[i] - np.nanmean(yobs[i]))**2)
            try:
                nash_dict[aba] = 1 - num/den
            except Exception as e:
                nash_dict[aba] = minus_inf
            if nash_dict[aba] < 0:
                negative_count += 1
            nash_sum += nash_dict[aba]
        avnse = nash_sum/n
        return avnse, nash_dict, negative_count

    # spatial nash-sutcliffe efficiency (SPATNSE)
    def calc_spatnse(self, sim_data, field_data):
        ndata = self.st.SWMM_Nperiods
        yobs = field_data.reshape(int(len(field_data)/ndata), ndata)
        ysim = sim_data.reshape(int(len(sim_data)/ndata), ndata)
        nash_dict = {}
        wb = openpyxl.load_workbook(self.data_filename)
        abas = wb.sheetnames
        negative_count = 0
        n = len(abas)
        num_sum = 0
        den_sum = 0
        sum_avg_yobs = 0
        for i,aba in enumerate(abas):
            sum_avg_yobs += np.nanmean(yobs[i])
        for i,aba in enumerate(abas):
            mask = ~np.isnan(yobs[i])
            avg_yobs = np.mean(yobs[i][mask])
            avg_ysim = np.mean(ysim[i][mask])
            num_sum += (avg_yobs - avg_ysim)**2
            den_sum += (avg_yobs - (1/n)*sum_avg_yobs)**2
            num = np.nansum((yobs[i] - ysim[i])**2)
            den = np.nansum((yobs[i] - np.nanmean(yobs[i]))**2)
            try:
                nash_dict[aba] = 1 - num/den
            except Exception as e:
                nash_dict[aba] = minus_inf
            if nash_dict[aba] < 0:
                negative_count += 1
        spatnse = 1 - num_sum/den_sum
        return spatnse, nash_dict, negative_count

    # regional nash-sutcliffe efficiency (REGNSE)
    def calc_regnse(self, sim_data, field_data):
        ndata = self.st.SWMM_Nperiods
        yobs = field_data.reshape(int(len(field_data)/ndata), ndata)
        ysim = sim_data.reshape(int(len(sim_data)/ndata), ndata)
        yobs_all = field_data
        ysim_all = sim_data
        nash_dict = {}
        wb = openpyxl.load_workbook(self.data_filename)
        abas = wb.sheetnames
        negative_count = 0
        n = len(abas)

        for i,aba in enumerate(abas):
            num = np.nansum((yobs[i] - ysim[i])**2)
            den = np.nansum((yobs[i] - np.nanmean(yobs[i]))**2)
            try:
                # print(num)
                # print(den)
                # pause()
                nash_dict[aba] = 1 - num/den
            except Exception as e:
                nash_dict[aba] = minus_inf
            if nash_dict[aba] < 0:
                negative_count += 1

        num_sum = np.nansum((yobs_all - ysim_all)**2)
        yobs_sum = np.nansum(yobs_all)
        yobs_size = np.count_nonzero(~np.isnan(yobs_all))
        den_sum = np.nansum((yobs_all - (1/yobs_size) * yobs_sum)**2)
        regnse = 1 - num_sum/den_sum
        return regnse, nash_dict, negative_count


    def calc_nash(self, of):
        try:
            sim_data = self.load_simulation_data()
            field_data = self.load_field_data()
        except Exception as e:
            raise
        if of == 'fo1':
            nse, dict, nc = self.calc_avnse(sim_data, field_data)
        if of == 'fo2':
            nse, dict, nc = self.calc_spatnse(sim_data, field_data)
        if of == 'fo3':
            nse, dict, nc = self.calc_regnse(sim_data, field_data)
        self.st.clean()
        return nse, dict, nc

    def run_swmm(self):
        if self.working_dir != None:
            os.chdir(self.working_dir)
        try:
            self.st = SWMM5Simulation(self.swmm_filename) # run a simulation
            return True
        except Exception as e:
            print(str(e))
            if str(e).strip() == '303:':
                print (str(e) + ' Arquivo {} nao encontrado.'.format(self.swmm_filename))
            if str(e).strip() == '317:' or str(e).strip() == '313:':
                print(str(e) + ' Erro ao carregar dados {}.'.format(self.swmm_filename))
            if self.temp_file:
                os.remove(self.swmm_filename)
            sys.exit()
        finally:
            if self.working_dir != None:
                os.chdir(self.root_dir)

def main():
    spreadsheet_name = 'data/dados-calib.xlsx'
    swmm_filename = 'final.inp'
    tmp_dir = 'tmp'

    # cria o objeto
    calib = SwmmTools(swmm_filename, spreadsheet_name, 'results')

    # # roda o swmm
    #calib.run_swmm()
    # print('Running swmm...')

    # # exibe os nos que possuem dados (baseado nas abas do excel)
    # nodes = calib.get_observed_nodes()
    # print(nodes)
    #
    # # pega os niveis simulados de um node especifico
    # niveis = calib.get_level_by_id('N-3')
    # print(niveis)
    #
    # # carrega os valores observados dos nodes
    # field_data = calib.load_field_data()
    # for data in field_data:
    #     print(data)

    # carrega os valores simulados dos nodes
    # sim_data = calib.load_simulation_data()
    # print(sim_data)
    #
    nash, dict, nc = calib.calc_nash()
    print('nash: ' + str(nash))
    print('dict: ' + str(dict))

if __name__ == '__main__':
    main()
