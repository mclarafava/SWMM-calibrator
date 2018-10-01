import openpyxl
import numpy as np
from swmm5.swmm5tools import SWMM5Simulation
import xlrd

from bokeh.io import output_file, show
from bokeh.layouts import gridplot
from bokeh.palettes import Viridis3
from bokeh.plotting import figure, ColumnDataSource
from bokeh.models import HoverTool
from datetime import datetime, timedelta
from collections import OrderedDict
import time

def swmm_plot():
    output_file("layout.html")
    spreadsheet_name = 'data/dados-calib.xlsx'
    swmm_filename = 'results/partial.inp'

    st=SWMM5Simulation(swmm_filename) # run a simulation

    wb = openpyxl.load_workbook(spreadsheet_name)
    abas = wb.sheetnames

    observado_todos = np.array([],dtype=np.float)
    simulado_todos = np.array([],dtype=np.float)

    for aba in abas:
        sheet = wb[aba]
        sheet.cell(row=1, column=4).value = 'Nivel_simulado'

        # lendo o nivel medido
        nivel_medido = np.array([], dtype=np.float)
        for i in range(2, sheet.max_row+1):
            nivel_medido = np.append(nivel_medido, sheet.cell(row=i,column=3).value)

        # pegando o simulado do swmm
        nivel_simulado = np.array(list(st.Results('NODE', aba, 0)))

        # escrevendo o simulado no excel
        ndados = st.SWMM_Nperiods
        for i in range(2, np.size(nivel_simulado)+2):
            sheet.cell(row=i, column=4).value = nivel_simulado[i-2]

        observado_todos = np.concatenate ((observado_todos, nivel_medido))
        simulado_todos = np.concatenate ((simulado_todos, nivel_simulado))

    data = np.array([], dtype=np.float)
    hora = np.array([], dtype=np.float)
    datahora = np.array([], dtype=np.datetime64)
    for i in range(2, sheet.max_row+1):
        data = np.append(data, sheet.cell(row=i,column=1).value)
        hora = np.append(hora, sheet.cell(row=i,column=2).value)
        #print(sheet.cell(row=i,column=2).value)
        diazin = sheet.cell(row=i,column=1).value
        horinha = sheet.cell(row=i,column=2).value
        datahora = np.append(datahora, diazin + timedelta(seconds=60*60*horinha.hour + 60*horinha.minute + horinha.second))

    observado_todos = np.array(observado_todos, dtype=np.float)
    simulado_todos = simulado_todos*100

    wb.save(spreadsheet_name)


    num = np.nansum((observado_todos - simulado_todos)**2)
    den = np.nansum((observado_todos - np.nanmean(observado_todos))**2)

    numlog = np.nansum((np.log(observado_todos) - np.log(simulado_todos))**2)
    denlog = np.nansum((np.log(observado_todos) - np.log(np.nanmean(observado_todos)))**2)

    if den == 0:
        den = 0.0000001
    if denlog == 0:
        denlog = 0.0000001

    print('Nash: %.4f' % (1-num/den))
    print('Log-Nash: %4f' % (1-numlog/denlog))

    observado_todos = observado_todos.reshape(int(len(observado_todos)/ndados), ndados)
    simulado_todos = simulado_todos.reshape(int(len(simulado_todos)/ndados), ndados)
    simulado_todos = simulado_todos/100
    x = datahora
    yobs = observado_todos
    ysim = simulado_todos

    nash = {}
    nashlog = {}
    for i,aba in enumerate(abas):
        num = np.nansum((yobs[i] - ysim[i])**2)
        den = np.nansum((yobs[i] - np.nanmean(yobs[i]))**2)
        numlog = np.nansum((np.log(yobs[i]) - np.log(ysim[i]))**2)
        denlog = np.nansum((np.log(yobs[i]) - np.log(np.nanmean(yobs[i])))**2)
        if den == 0:
            den = 0.0000001
        if denlog == 0:
            denlog = 0.0000001
        nash[aba] = 1 - num/den
        nashlog[aba] = 1 - numlog/denlog

    figuras = {}
    headers = {}
    for i, aba in enumerate(abas):
        headers[aba] = 'Nash: %.4f' % nash[aba] + ' | ' + 'Log-Nash: %.4f' % nashlog[aba]
        figuras[aba] = figure(width=2100, plot_height=500,
                               title=aba + ' | '+headers[aba], x_axis_type="datetime")
        figuras[aba].title.text_font_size = '16pt'
        figuras[aba].xaxis.axis_label_text_font_size = '14pt'
        figuras[aba].yaxis. axis_label_text_font_size = '14pt'
        figuras[aba].circle(x, yobs[i], fill_color=None, line_color="navy")
        figuras[aba].line(x, yobs[i], color='navy')
        figuras[aba].line(x, ysim[i], color='firebrick')

    figs = [[figuras[figura]] for figura in abas]
    grid = gridplot(figs)
    show(grid)
swmm_plot()
